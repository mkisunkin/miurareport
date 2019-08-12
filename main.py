from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from testxls import wb, ft, fill, al
from testxls import Table
import testxls
from report import patient_list, num_pat, latest_file
from other_const import table_symbol, tests_dick, path_report
import datetime
import os
import glob


table = Table()

wb = Workbook()
ws0 = wb.active


'''Проверка на наличие папки для отчета'''

if not os.path.isdir(path_report):
    os.makedirs(path_report)


print('Введите имя врача: ')
doc_name = input()
date_test = datetime.datetime.today()

'''Заполнение первой страницы'''

ws0.merge_cells('A1:I1')
ws0['A1'] = 'Лабораторный отчет'
ws0['A1'].font = ft[1]
ws0['A1'].alignment = al[1]
ws0.merge_cells('H2:I2')
ws0['H2'] = date_test.strftime("%d/%m/%Y_ %H:%M")

'''Отрисовка первой страницы'''

count_pat = len(patient_list)
step_row = 0


for i in range(count_pat):
    step_row = step_row + 4
    step_row_str = str(step_row)
    step_row_str_res = str(step_row + 1 )
    step_row_str_name = str(step_row - 1)
    ws0['A' + step_row_str_name] = patient_list[i].name
    ws0['D' + step_row_str_name] = 'Prob ' + patient_list[i].num_prob
    yrow = len(patient_list[i].tests)
    ''' В одну строчку убирается  9 тестов '''
    if yrow <= 9: # Если количество тестов меньше 9
        for k in range(yrow):
            xcol_tests = str(table_symbol[k] + step_row_str)
            xcol_results = str(table_symbol[k] + step_row_str_res)
            ws0[xcol_tests] = patient_list[i].tests[k]
            ws0[xcol_tests].fill = fill
            ws0[xcol_results] = patient_list[i].results[k]
    else:  # Если количество тестов больше 9
        for k in range(9):
            xcol_tests = str(table_symbol[k] + step_row_str)
            xcol_results = str(table_symbol[k] + step_row_str_res)
            ws0[xcol_tests] = patient_list[i].tests[k]
            ws0[xcol_tests].fill = fill
            ws0[xcol_results] = patient_list[i].results[k]
        hi_yrow = yrow - 9
        step_row = step_row + 3
        step_row_str = str(step_row)
        step_row_str_res = str(step_row + 1)
        for k in range(hi_yrow):
            xcol_tests = str(table_symbol[k] + step_row_str)
            xcol_results = str(table_symbol[k] + step_row_str_res)
            ws0[xcol_tests] = patient_list[i].tests[k]
            ws0[xcol_tests].fill = fill
            ws0[xcol_results] = patient_list[i].results[k]


sheets = ['']

''' Цикл отрисовывающий книги под каждого пациента'''
for i in range(count_pat):
    k = str(patient_list[i].num_prob)
    
    
    footer_row = str(16 + len(patient_list[i].tests))
    patient_name_sheet = 0
    patient_name_sheet = wb.create_sheet('Prob ' + k)
    table.cells_merge(patient_name_sheet)
    table.add_image(patient_name_sheet)
    table.add_standart_text( patient_name_sheet)
    table.create_border(patient_name_sheet)
    table.table_border( patient_name_sheet, str(14))

    patient_name_sheet['B9'] = patient_list[i].name
    stp_row = 14
    table.footer(patient_name_sheet, footer_row)
    patient_name_sheet['D' + footer_row ] = date_test.strftime("%d/%m/%Y")
    patient_name_sheet['H' + footer_row] = doc_name
    
    for j in range(len(patient_list[i].tests)):
        stp_row += 1
        stp_row_str = str(stp_row )
        table.table_merge(patient_name_sheet, stp_row_str)
        table.table_border(patient_name_sheet, stp_row_str)
        patient_name_sheet['A' + stp_row_str] = tests_dick[patient_list[i].tests[j]][0]
        patient_name_sheet['A' + stp_row_str].font = ft[2]
        patient_name_sheet['E' + stp_row_str] = patient_list[i].results[j]
        patient_name_sheet['E' + stp_row_str].font = ft[2]
        patient_name_sheet['G' + stp_row_str] = tests_dick[patient_list[i].tests[j]][1]
        patient_name_sheet['G' + stp_row_str].font = ft[2]
        patient_name_sheet['H' + stp_row_str] = tests_dick[patient_list[i].tests[j]][2]
        patient_name_sheet['H' + stp_row_str].font = ft[2]
        if j % 2 != 0: #заполнение ячеек цветом
            patient_name_sheet['A' + stp_row_str].fill = fill
            patient_name_sheet['E' + stp_row_str].fill = fill
            patient_name_sheet['G' + stp_row_str].fill = fill
            patient_name_sheet['H' + stp_row_str].fill = fill
        else : pass

    sheets.append(patient_name_sheet)

file_name = latest_file[23:-3]
report_file_name = 'rep/' + file_name + 'xlsx'
print(report_file_name)
print('Готово')

wb.save(report_file_name)
