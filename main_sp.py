from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from testxls import wb, ft, fill, al
from testxls import Table
import testxls
from report import patient_list, num_pat, latest_file
from other_const import table_symbol, tests_dick, path_report
import datetime
import os
import glob

table = Table()

#wb = Workbook()
#ws0 = wb.active

if not os.path.isdir(path_report):
    os.makedirs(path_report)

print('Введите имя врача: ')
doc_name = input()
date_test = datetime.datetime.today()

