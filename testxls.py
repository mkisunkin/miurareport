from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import report
from report import patient_list

wb = Workbook()
ws0 = wb.active

''' Основные параметры и константы'''

ft = list(range(6))
al = list(range(5))
ft[0] = 0
al[0] = 0

ft[1] = Font(name='Arial', size=14, bold=True, italic=True)
ft[2] = Font(name='Arial', size=10, bold=False, italic=True)
ft[3] = Font(name='Arial', size=11, bold=False, italic=False)
ft[4] = Font(name='Arial', size=11, bold=True, italic=False)
ft[5] = Font(name='Arial', size=10, bold=True, italic=False)

al[1] = Alignment(horizontal= 'center')
al[2] = Alignment(horizontal='right')
al[3] = Alignment(horizontal='center', vertical= 'center')
al[4] = Alignment(horizontal='center', vertical='center', wrap_text=True)

bord_hor = Border(bottom=Side(border_style='thin', color='000000'))
bord_top = Border(top=Side(border_style='thin', color='000000'))
bord_hor_left = Border(left=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'))
bord_hor_right = Border(right=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'))
bord_top_right = Border(right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'))
bord_top_left = Border(left=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'))

fill = PatternFill("solid", fgColor="DDDDDD")
<<<<<<< HEAD
#fill = PatternFill(fill_type=None, start_color='FFFFFFFF',end_color='FF000000')


'''Объединение ячеек'''
def cells_merge(name):
    ws = name
    ws.merge_cells('A6:I6')
    ws.merge_cells('A7:I7')
    ws.merge_cells('B9:D9')
    ws.merge_cells('B10:D10')
    ws.merge_cells('B11:D11')

    ws.merge_cells('E9:F9')
    ws.merge_cells('E10:F10')
    ws.merge_cells('E11:F11')

    ws.merge_cells('G9:I9')
    ws.merge_cells('G10:I10')
    ws.merge_cells('G11:I11')

    ws.merge_cells('A13:D14')
    ws.merge_cells('E13:F14')
    ws.merge_cells('G13:G14')
    ws.merge_cells('H13:I14')

    ws.merge_cells('A47:C47')
    ws.merge_cells('D47:E47')
    ws.merge_cells('F47:G47')
    ws.merge_cells('H47:I47')

    diap1 = ' '
    diap2 = ' '
    diap3 = ' '


    for i in range(15,46):
        k = str(i)
        diap1 = str('A'+ k + ':D'+ k)
        diap2 = str('E' + k + ':F' + k)
        diap3 = str('H' + k + ':I' + k)
        ws.merge_cells(diap1)
        ws.merge_cells(diap2)
        ws.merge_cells(diap3)

'''Изображение'''
def create_image(name):
    ws = name
    img = Image('image002.png')
    ws.add_image(img, 'A1')

''' создание рамки таблицы'''
def border_table(rows_table):
        rows_table_str = str(rows_table)
        


'''Рамка'''

def create_border(name):
    ws = name
    ws['B9'].border = bord_hor
    ws['B10'].border = bord_hor
    ws['B11'].border = bord_hor

    ws['C9'].border = bord_hor
    ws['C10'].border = bord_hor
    ws['C11'].border = bord_hor

    ws['D9'].border = bord_hor
    ws['D10'].border = bord_hor
    ws['D11'].border = bord_hor


    ws['G9'].border = bord_hor
    ws['G10'].border = bord_hor
    ws['G11'].border = bord_hor

    ws['H9'].border = bord_hor
    ws['H10'].border = bord_hor
    ws['H11'].border = bord_hor

    ws['I9'].border = bord_hor
    ws['I10'].border = bord_hor
    ws['I11'].border = bord_hor


'''Текст'''
def add_standart_text(name):
    ws = name
    ws['A6'] = 'Биохимический анализ крови'
    ws['A7'] = 'автоматический анализатор Miura, I.C.E. group'
    a6 = ws['A6']
    a6.font = ft[1]
    a6.alignment = al[1]

    a7 = ws['A7']
    a7.font = ft[2]
    a7.alignment = al[1]

    ws['A9'] = 'Ф.И.О'
    ws['A9'].font = ft[3]
    ws['A10'] = 'Врач'
    ws['A10'].font = ft[3]
    ws['A11'] = 'Диагноз'
    ws['A11'].font = ft[3]
    ws['G10'] = 'МЦ "Ультрамед" '
    ws['G10'].font = ft[3]
    ws['G10'].alignment = al[1]

    ws['E9'] = 'Возраст'
    ws['E9'].font = ft[3]
    ws['E9'].alignment = al[2]
    ws['E10'] = 'ЛПУ'
    ws['E10'].font = ft[3]
    ws['E10'].alignment = al[2]
    ws['E11'] = 'Дата забора'
    ws['E11'].font = ft[3]
    ws['E11'].alignment = al[2]

    ws['A13'] = 'Показатель'
    a13 = ws['A13']
    a13.alignment = al[3]
    a13.font = ft[4]
    ws['A13'].fill = fill

    ws['E13'] = 'Результат'
    ws['E13'].alignment = al[3]
    ws['E13'].font = ft[4]
    ws['E13'].fill = fill

    ws['G13'] = 'Ед. изм.'
    ws['G13'].alignment = al[3]
    ws['G13'].font = ft[4]
    ws['G13'].fill = fill

    ws['H13'] = 'Референсные значения'
    ws['H13'].alignment = al[4]
    ws['H13'].font = ft[4]
    ws['H13'].fill = fill

    ws['A47'] = 'Дата выдачи результата:'
    ws['A47'].alignment = al[2]
    ws['A47'].font = ft[5]

    ws['F47'] = 'Выполнил:'
    ws['F47'].alignment = al[2]
    ws['F47'].font = ft[5]

    ws['B10'] = 'амб'
    ws['B11'] = 'Обследование'



''' Функция отрисовки таблицы'''
def create_table(name_pat):
    
    cells_merge(name_pat)
    create_border(name_pat)
    create_image(name_pat)
    add_standart_text(name_pat)
=======
>>>>>>> 593c4f621386bf8dc273833ec4bf885be1b75ca1

class Table():
        def add_image(self,ws): # добавление лого
                img = Image('image002.png')
                ws.add_image(img, 'A1')

        def footer(self, ws, rows_footer): #отрисовка футера

                ws.merge_cells('A' + rows_footer + ':C' + rows_footer)
                ws.merge_cells('D' + rows_footer + ':E' + rows_footer)
                ws.merge_cells('F' + rows_footer + ':G' + rows_footer)
                ws.merge_cells('H' + rows_footer + ':I' + rows_footer)

                ws['A' + rows_footer] = 'Дата выдачи результата:'
                ws['A' + rows_footer].alignment = al[2]
                ws['A' + rows_footer].font = ft[5]
                ws['F' + rows_footer] = 'Выполнил:'
                ws['F' + rows_footer].alignment = al[2]
                ws['F' + rows_footer].font = ft[5]

        def add_standart_text(self,ws): #отрисовка статичного текста
                ws['A6'] = 'Биохимический анализ крови'
                ws['A7'] = 'автоматический анализатор Miura, I.C.E. group'
                a6 = ws['A6']
                a6.font = ft[1]
                a6.alignment = al[1]

                a7 = ws['A7']
                a7.font = ft[2]
                a7.alignment = al[1]

                ws['A9'] = 'Ф.И.О'
                ws['A9'].font = ft[3]
                ws['A10'] = 'Врач'
                ws['A10'].font = ft[3]
                ws['A11'] = 'Диагноз'
                ws['A11'].font = ft[3]
                ws['G10'] = 'МЦ "Ультрамед" '
                ws['G10'].font = ft[3]
                ws['G10'].alignment = al[1]

                ws['E9'] = 'Возраст'
                ws['E9'].font = ft[3]
                ws['E9'].alignment = al[2]
                ws['E10'] = 'ЛПУ'
                ws['E10'].font = ft[3]
                ws['E10'].alignment = al[2]
                ws['E11'] = 'Дата забора'
                ws['E11'].font = ft[3]
                ws['E11'].alignment = al[2]

                ws['A13'] = 'Показатель'
                ws['A13'].alignment = al[3]
                ws['A13'].font = ft[4]
                ws['A13'].fill = fill

                ws['E13'] = 'Результат'
                ws['E13'].alignment = al[3]
                ws['E13'].font = ft[4]
                ws['E13'].fill = fill

                ws['G13'] = 'Ед. изм.'
                ws['G13'].alignment = al[3]
                ws['G13'].font = ft[4]
                ws['G13'].fill = fill

                ws['H13'] = 'Референсные значения'
                ws['H13'].alignment = al[4]
                ws['H13'].font = ft[4]
                ws['H13'].fill = fill

                ws['B10'] = 'амб'
                ws['B11'] = 'Обследование'

        def create_border(self,ws):
                ws['B9'].border = bord_hor
                ws['B10'].border = bord_hor
                ws['B11'].border = bord_hor

                ws['C9'].border = bord_hor
                ws['C10'].border = bord_hor
                ws['C11'].border = bord_hor

                ws['D9'].border = bord_hor
                ws['D10'].border = bord_hor
                ws['D11'].border = bord_hor

                ws['G9'].border = bord_hor
                ws['G10'].border = bord_hor
                ws['G11'].border = bord_hor

                ws['H9'].border = bord_hor
                ws['H10'].border = bord_hor
                ws['H11'].border = bord_hor

                ws['I9'].border = bord_hor
                ws['I10'].border = bord_hor
                ws['I11'].border = bord_hor

                #table_border(ws, str(14))

                ws['A13'].border = bord_top_left
                ws['B13'].border = bord_top
                ws['C13'].border = bord_top
                ws['D13'].border = bord_top
                ws['E13'].border = bord_top_left
                ws['F13'].border = bord_top
                ws['G13'].border = bord_top_left
                ws['H13'].border = bord_top_left
                ws['I13'].border = bord_top_right

        def cells_merge(self,ws):
                ws.merge_cells('A6:I6')
                ws.merge_cells('A7:I7')
                ws.merge_cells('B9:D9')
                ws.merge_cells('B10:D10')
                ws.merge_cells('B11:D11')

                ws.merge_cells('E9:F9')
                ws.merge_cells('E10:F10')
                ws.merge_cells('E11:F11')

                ws.merge_cells('G9:I9')
                ws.merge_cells('G10:I10')
                ws.merge_cells('G11:I11')

                ws.merge_cells('A13:D14')
                ws.merge_cells('E13:F14')
                ws.merge_cells('G13:G14')
                ws.merge_cells('H13:I14')

        def table_border(self, ws, rows_table):
                ws['A' + rows_table].border = bord_hor_left
                ws['B' + rows_table].border = bord_hor
                ws['C' + rows_table].border = bord_hor
                ws['D' + rows_table].border = bord_hor
                ws['E' + rows_table].border = bord_hor_left
                ws['F' + rows_table].border = bord_hor
                ws['G' + rows_table].border = bord_hor_left
                ws['H' + rows_table].border = bord_hor_left
                ws['I' + rows_table].border = bord_hor_right

        def table_merge(self,ws, rows_table):
                ws.merge_cells('A' + rows_table + ':D' + rows_table)
                ws.merge_cells('E' + rows_table + ':F' + rows_table)
                ws.merge_cells('H' + rows_table + ':I' + rows_table)
